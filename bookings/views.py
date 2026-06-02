import csv
import json
from django.http import HttpResponse
import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone as dj_timezone
from django.contrib.auth.models import Group
from .models import Booking, StudioSchedule, StudioException, BookingFile, GroupMember, Equipment
from .forms import BookingForm
from accounts.models import CustomUser
from accounts.forms import CustomUserChangeForm
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib.admin.views.decorators import staff_member_required

def home(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'about.html')

def manual(request):
    return render(request, 'manual.html')

@login_required
def profile(request):
    return render(request, 'profile.html')

@login_required
def profile_edit(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            return redirect('profile')
    else:
        form = CustomUserChangeForm(instance=request.user)
    return render(request, 'profile_edit.html', {'form': form})

@login_required
def notifications(request):
    # Показываем последние изменения бронирований пользователя (отсортированы по убыванию updated_at, если добавим поле)
    # В качестве уведомлений можно взять все брони пользователя, отсортированные по статусу и дате.
    user_bookings = Booking.objects.filter(user=request.user).order_by('-created_at')[:20]
    return render(request, 'notifications.html', {'notifications': user_bookings})

@login_required
def get_calendar_data(request):
    """Возвращает JSON с днями и слотами для FullCalendar."""
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    if not start_str or not end_str:
        return JsonResponse({'error': 'Missing start/end'}, status=400)

    try:
        start_date = datetime.datetime.strptime(start_str.split('T')[0], '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_str.split('T')[0], '%Y-%m-%d').date()
    except (ValueError, IndexError):
        return JsonResponse({'error': 'Invalid date format'}, status=400)

    schedules = StudioSchedule.objects.all()
    exceptions = StudioException.objects.filter(date__gte=start_date, date__lte=end_date)
    bookings = Booking.objects.filter(
        status__in=['booked', 'completed'],
        start_time__date__gte=start_date,
        end_time__date__lte=end_date + datetime.timedelta(days=1)
    ).select_related('user')

    teachers_group = Group.objects.get(name='Teachers')

    days = []
    current = start_date
    now = dj_timezone.now()  # текущий момент один раз для всего запроса
    while current <= end_date:
        dow = current.isoweekday()
        # Сначала ищем расписание для конкретной даты (спец. день)
        day_schedule = schedules.filter(specific_date=current).first()
        # Если спец. дня нет — используем обычное расписание по дню недели
        if not day_schedule:
            day_schedule = schedules.filter(day_of_week=dow, specific_date__isnull=True).first()
        day_exceptions = [e for e in exceptions if e.date == current]

        if not day_schedule:
            days.append({
                'date': current.isoformat(),
                'color': 'red',
                'slots': []
            })
            current += datetime.timedelta(days=1)
            continue

        start_hour = day_schedule.start_time.hour
        end_hour = day_schedule.end_time.hour
        slots = []
        any_available = False
        any_unavailable = False

        for hour in range(start_hour, end_hour):
            slot_start = dj_timezone.make_aware(
                datetime.datetime.combine(current, datetime.datetime.min.time().replace(hour=hour))
            )
            slot_end = slot_start + datetime.timedelta(hours=1)
            slot_available = True
            occupied_by = None

            # Проверка, не прошёл ли слот (время старта меньше текущего)
            if slot_start <= now:
                slot_available = False
                # occupied_by оставляем None, чтобы слот был просто серым
            else:
                # Проверка исключений
                for exc in day_exceptions:
                    exc_start = dj_timezone.make_aware(datetime.datetime.combine(current, exc.start_time))
                    exc_end = dj_timezone.make_aware(datetime.datetime.combine(current, exc.end_time))
                    if slot_start < exc_end and slot_end > exc_start:
                        slot_available = False
                        occupied_by = 'exception'
                        break

                # Проверка броней
                if slot_available:
                    slot_bookings = [
                        b for b in bookings
                        if b.start_time < slot_end and b.end_time > slot_start
                    ]
                    if slot_bookings:
                        slot_available = False
                        if any(b.user.groups.filter(name='Teachers').exists() for b in slot_bookings):
                            occupied_by = 'teacher'
                        else:
                            occupied_by = 'student'

            # Детали для преподавателя
            details = ''
            if occupied_by == 'student' and request.user.groups.filter(name='Teachers').exists():
                occupant = slot_bookings[0].user.email if slot_bookings else ''
                details = f'Занято: {occupant}'

            slots.append({
                'time': f'{hour:02d}:00 – {hour+1:02d}:00',
                'start': slot_start.isoformat(),
                'available': slot_available,
                'occupied_by': occupied_by,
                'details': details,
            })

            if slot_available:
                any_available = True
            else:
                any_unavailable = True

        if not any_available and any_unavailable:
            color = 'red'
        elif any_available and not any_unavailable:
            color = 'green'
        else:
            color = 'yellow'

        days.append({
            'date': current.isoformat(),
            'color': color,
            'slots': slots
        })
        current += datetime.timedelta(days=1)

    return JsonResponse({'days': days})

@login_required
def create_booking(request):
    if request.method == 'POST':
        start_str = request.POST.get('start_time')
        if not start_str:
            msg = 'Выберите время.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': msg})
            messages.error(request, msg)
            return redirect('bookings:create_booking')

        try:
            start = datetime.datetime.fromisoformat(start_str)
            if dj_timezone.is_naive(start):
                start = dj_timezone.make_aware(start)
        except (ValueError, TypeError):
            msg = 'Неверный формат времени.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': msg})
            messages.error(request, msg)
            return redirect('bookings:create_booking')

        end = start + datetime.timedelta(hours=1)
        description = request.POST.get('description', '')
        files = request.FILES.getlist('file')

        is_teacher = request.user.groups.filter(name='Teachers').exists()
        is_guest = request.user.groups.filter(name='ExternalGuests').exists()

        conflicts = Booking.objects.filter(
            status__in=['booked', 'completed'],
            start_time__lt=end,
            end_time__gt=start
        )

        if conflicts.exists():
            if not is_teacher:
                msg = 'Это время уже занято.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': msg})
                messages.error(request, msg)
                return redirect('bookings:create_booking')
            else:
                if conflicts.filter(user__groups__name='Teachers').exists():
                    msg = 'Слот занят другим преподавателем.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': msg})
                    messages.error(request, msg)
                    return redirect('bookings:create_booking')
                else:
                    for conflict in conflicts:
                        conflict.status = Booking.Status.CANCELLED
                        conflict.description += '\n(Отменено преподавателем)'
                        conflict.save()

        form = BookingForm({
            'start_time': start,
            'end_time': end,
            'description': description,
        })
        if form.is_valid():
            booking = form.save(commit=False)
            booking.user = request.user
            booking.save()


            for f in files:
                BookingFile.objects.create(booking=booking, file=f)

            # Привязка выбранного оборудования
            equipment_ids = request.POST.getlist('equipment')
            if equipment_ids:
                booking.equipment.set(equipment_ids)

            # Групповые участники
            members_json = request.POST.get('members')
            if members_json:
                try:
                    members = json.loads(members_json)
                    for member in members:
                        email = member.get('email', '').strip()
                        full_name = member.get('full_name', '')
                        phone = member.get('phone', '')
                        user = CustomUser.objects.filter(email=email).first()
                        if user:
                            user_booking = Booking.objects.create(
                                user=user,
                                start_time=booking.start_time,
                                end_time=booking.end_time,
                                status=Booking.Status.BOOKED,
                                description=booking.description,
                            )
                            for f in files:
                                BookingFile.objects.create(booking=user_booking, file=f)
                        GroupMember.objects.create(
                            booking=booking,
                            full_name=full_name,
                            email=email,
                            phone=phone
                        )
                except json.JSONDecodeError:
                    pass

            msg = 'Бронирование создано!'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': msg})
            messages.success(request, msg)
            return redirect('bookings:booking_list')
        else:
            errors = [err for errs in form.errors.values() for err in errs]
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': ' '.join(errors)})
            for err in errors:
                messages.error(request, err)
            today = dj_timezone.now().date()
            return render(request, 'bookings/booking_form.html', {'today': today})

    # GET
    today = dj_timezone.now().date()
    is_teacher = request.user.groups.filter(name='Teachers').exists()
    is_admin = request.user.is_staff  # ← новая строка
    all_equipment = Equipment.objects.all()
    return render(request, 'bookings/booking_form.html', {
        'today': today,
        'is_teacher': is_teacher,
        'is_admin': is_admin,          # ← передаём в шаблон
        'all_equipment': all_equipment,
    })

@login_required
def booking_list(request):
    bookings = Booking.objects.filter(user=request.user).order_by('-start_time')
    all_equipment = Equipment.objects.all()
    return render(request, 'bookings/booking_list.html', {
        'bookings': bookings,
        'all_equipment': all_equipment,
    })

@login_required
def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    if booking.status == Booking.Status.BOOKED:
        booking.status = Booking.Status.CANCELLED
        booking.save()
        messages.success(request, 'Бронь отменена.')
    else:
        messages.error(request, 'Нельзя отменить эту бронь.')
    return redirect('bookings:booking_list')

def get_slots(request):
    """Возвращает доступные слоты на выбранную дату в JSON, учитывая роли."""
    import datetime as dt
    from django.utils import timezone
    from .models import StudioSchedule, StudioException, Booking
    from django.contrib.auth.models import Group

    date_str = request.GET.get('date')
    if not date_str:
        return JsonResponse({'slots': []})

    try:
        date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'slots': []})

    day_of_week = date.isoweekday()
    schedule_qs = StudioSchedule.objects.filter(day_of_week=day_of_week)
    if not schedule_qs:
        return JsonResponse({'slots': []})

    schedule = schedule_qs.first()
    start_hour = schedule.start_time.hour
    end_hour = schedule.end_time.hour

    exceptions = StudioException.objects.filter(date=date)
    slots = []
    now = timezone.now()

    # Заранее определим группы, чтобы избежать запросов в цикле
    teachers_group = Group.objects.get(name='Teachers')

    for hour in range(start_hour, end_hour):
        slot_start = timezone.make_aware(dt.datetime.combine(date, dt.time(hour, 0)))
        slot_end = slot_start + dt.timedelta(hours=1)

        available = True
        occupied_by = None  # 'student' или 'teacher'

        if slot_start <= now:
            available = False
        else:
            # Проверка исключений
            for exc in exceptions:
                exc_start = timezone.make_aware(dt.datetime.combine(date, exc.start_time))
                exc_end = timezone.make_aware(dt.datetime.combine(date, exc.end_time))
                if slot_start < exc_end and slot_end > exc_start:
                    available = False
                    break

            # Проверка занятости
            if available:
                conflict_bookings = Booking.objects.filter(
                    status__in=['booked', 'completed'],
                    start_time__lt=slot_end,
                    end_time__gt=slot_start
                )
                if conflict_bookings.exists():
                    # Проверяем, есть ли среди конфликтующих броней преподавательская
                    if conflict_bookings.filter(user__groups=teachers_group).exists():
                        available = False
                        occupied_by = 'teacher'
                    else:
                        available = False
                        occupied_by = 'student'  # занято студентом/гостем

        slots.append({
            'time': f'{hour:02d}:00 – {hour+1:02d}:00',
            'start': slot_start.isoformat(),
            'available': available,
            'occupied_by': occupied_by,
        })

    return JsonResponse({'slots': slots})

@login_required
def edit_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id, user=request.user)
    if booking.status != Booking.Status.BOOKED:
        return JsonResponse({'success': False, 'error': 'Редактирование разрешено только для активных броней.'}, status=403)

    if request.method == 'POST':
        description = request.POST.get('description', '')
        booking.description = description
        booking.save()

        # Удаление файлов
        delete_files = request.POST.getlist('delete_files')
        if delete_files:
            BookingFile.objects.filter(id__in=delete_files, booking=booking).delete()

        # Новые файлы
        new_files = request.FILES.getlist('files')
        for f in new_files:
            BookingFile.objects.create(booking=booking, file=f)

        # Управление участниками
        delete_members_ids = request.POST.getlist('delete_members')
        if delete_members_ids:
            GroupMember.objects.filter(id__in=delete_members_ids, booking=booking).delete()

        members_json = request.POST.get('members')
        if members_json:
            try:
                members = json.loads(members_json)
                for member in members:
                    member_id = member.get('id')
                    if member_id:
                        # Обновляем существующего участника
                        try:
                            gm = GroupMember.objects.get(id=member_id, booking=booking)
                            gm.full_name = member.get('full_name', '')
                            gm.email = member.get('email', '')
                            gm.phone = member.get('phone', '')
                            gm.save()
                        except GroupMember.DoesNotExist:
                            pass
                    else:
                        # Новый участник
                        GroupMember.objects.create(
                            booking=booking,
                            full_name=member.get('full_name', ''),
                            email=member.get('email', ''),
                            phone=member.get('phone', '')
                        )
            except json.JSONDecodeError:
                pass
        equipment_ids = request.POST.getlist('equipment')
        booking.equipment.set(equipment_ids)        
        return JsonResponse({'success': True, 'message': 'Бронирование обновлено.'})

    # GET – отдаём данные для модального окна
    files_data = [{'id': f.id, 'name': f.file.name, 'url': f.file.url} for f in booking.files.all()]
    members_data = [{'id': m.id, 'full_name': m.full_name, 'email': m.email, 'phone': m.phone} for m in booking.group_members.all()]
    equipment_ids = list(booking.equipment.values_list('id', flat=True))
    return JsonResponse({
        'description': booking.description,
        'files': files_data,
        'members': members_data,
        'equipment': equipment_ids,
    })

def logout_view(request):
    auth_logout(request)
    return redirect('home')

def is_teacher_or_admin(user):
    return user.is_staff or user.groups.filter(name='Teachers').exists()

@user_passes_test(is_teacher_or_admin, login_url='/accounts/login/')
def schedule_view(request):
    # Показываем завершённые (прошедшие) и будущие активные брони
    bookings = Booking.objects.filter(
        status__in=['completed', 'booked']
    ).select_related('user').prefetch_related('equipment').order_by('-start_time')
    
    return render(request, 'bookings/schedule.html', {'bookings': bookings})

@user_passes_test(is_teacher_or_admin, login_url='/accounts/login/')
def schedule_data(request):
    # Этот эндпоинт больше не нужен, но оставим заглушку для совместимости
    return JsonResponse({'bookings': []})

@staff_member_required
def export_bookings(request):
    fmt = request.GET.get('format', 'csv').lower()
    bookings = Booking.objects.select_related('user').prefetch_related(
        'files', 'group_members', 'equipment'
    ).all()

    # Подготовка данных в виде списка словарей
    data = []
    for b in bookings:
        data.append({
            'ID': b.id,
            'User': b.user.email,
            'Start': b.start_time.isoformat(),
            'End': b.end_time.isoformat(),
            'Description': b.description or '',
            'Status': b.status,
            'Created': b.created_at.isoformat(),
            'Files': ', '.join(f.file.name for f in b.files.all()),
            'Group Members': ', '.join(
                f'{m.full_name} ({m.email})' for m in b.group_members.all()
            ),
            'Equipment': ', '.join(eq.name for eq in b.equipment.all()),
        })

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="bookings.csv"'
        writer = csv.DictWriter(response, fieldnames=data[0].keys() if data else [])
        writer.writeheader()
        writer.writerows(data)
        return response

    elif fmt == 'json':
        response = HttpResponse(
            json.dumps(data, ensure_ascii=False, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="bookings.json"'
        return response

    elif fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse('openpyxl not installed', status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        if data:
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}1'] = header
            for row_num, row in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    ws[f'{col_letter}{row_num}'] = row[header]

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'
        wb.save(response)
        return response

    return HttpResponse('Unsupported format', status=400)